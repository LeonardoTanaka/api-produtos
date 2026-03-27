import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def gerar_excel(produtos):
    data = []

    for p in produtos:
        data.append({
            "nome": p["nome"],
            "preco": p["preco"]
        })

    df = pd.DataFrame(data)

    filename = f"produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    df.to_excel(filename, index=False)

    # 👉 aplicar formatação no Excel
    wb = load_workbook(filename)
    ws = wb.active

    # coluna B = preço
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    wb.save(filename)

    return filename